"""
Vistas para el dashboard principal
"""
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from .utils import association_required
from core.api import get_client
from dateutil import parser
import csv
import io
from socias.models import Socia
from finanzas.models import Transaccion
from eventos.models import Evento, Lugar
from proyectos.models import Proyecto
from entidades.models import Persona, Material


@association_required
def dashboard(request):
    """Dashboard principal de la asociación"""
    asociacion = request.user.profile.asociacion
    context = {
        'asociacion': asociacion,
        'user_profile': request.user.profile,
        'section': 'dashboard'
    }
    return render(request, 'dashboard/dashboard.html', context)


@association_required
def backend_management(request):
    """Vista unificada para gestión de backend (Datos y Drive)"""
    asociacion = request.user.profile.asociacion
    active_tab = request.GET.get('tab', 'database')

    # Contexto para Drive
    client = get_client(request)
    is_connected = False
    folder_link = None
    auth_url = None
    folders = []
    files = []

    try:
        # Verificar conexión
        config_status = client.get(f'drive/config/{asociacion.id}')
        is_connected = config_status.get('is_connected', False)
        folder_link = config_status.get('folder_link')

        if not is_connected:
            response = client.get('drive/auth/url')
            auth_url = response.get('url')
        else:
            # Obtener carpetas para configuración
            folders = client.get('drive/folders', params={'asociacion_id': asociacion.id})

            # Obtener archivos si hay carpeta configurada
            if asociacion.drive_folder_id:
                files = client.get('drive/files', params={'asociacion_id': asociacion.id})
                for f in files:
                    if f.get('createdTime'):
                        try:
                            f['createdTime'] = parser.parse(f['createdTime'])
                        except (ValueError, TypeError):
                            pass
    except Exception:
        pass

    return render(request, 'dashboard/backend_management.html', {
        'section': 'backend_management',
        'active_tab': active_tab,
        'asociacion': asociacion,
        'is_connected': is_connected,
        'auth_url': auth_url,
        'folder_link': folder_link,
        'folders': folders,
        'files': files
    })


@association_required
def drive_upload(request):
    """Procesa subida de archivos a Drive"""
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            client = get_client(request)
            asociacion = request.user.profile.asociacion
            uploaded_file = request.FILES['file']
            files = {'file': (uploaded_file.name, uploaded_file.read(), uploaded_file.content_type)}

            client.post('drive/upload',
                        data={'asociacion_id': asociacion.id},
                        files=files)
            messages.success(request, 'Archivo subido correctamente')
        except Exception as e:
            messages.error(request, f'Error al subir archivo: {str(e)}')

    return redirect('users:backend_management', permanent=False) # + '?tab=drive' handled by url params if needed, but redirect usually strips params unless added.
    # Better: return redirect('/dashboard/backend/?tab=drive') or use reverse with query params

@association_required
def drive_delete(request):
    """Procesa eliminación de archivos de Drive"""
    if request.method == 'POST' and request.POST.get('delete_file_id'):
        try:
            client = get_client(request)
            file_id = request.POST.get('delete_file_id')
            client.delete(f'drive/files/{file_id}')
            messages.success(request, 'Archivo eliminado correctamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar archivo: {str(e)}')

    return redirect('/dashboard/backend/?tab=drive')

@association_required
def drive_create_folder(request):
    """Procesa creación de carpeta en Drive"""
    if request.method == 'POST':
        try:
            client = get_client(request)
            asociacion = request.user.profile.asociacion
            folder_name = request.POST.get('folder_name', 'Gestor Asociaciones')

            response = client.post('drive/create-folder', data={
                'asociacion_id': asociacion.id,
                'folder_name': folder_name
            })

            if response and 'folder' in response:
                asociacion.drive_folder_id = response['folder']['id']
                asociacion.save()
                messages.success(request, f'Carpeta "{folder_name}" creada correctamente')
        except Exception as e:
            messages.error(request, f'Error al crear carpeta: {str(e)}')

    return redirect('/dashboard/backend/?tab=drive')

@association_required
def drive_select_folder(request):
    """Procesa selección de carpeta existente"""
    if request.method == 'POST':
        try:
            client = get_client(request)
            asociacion = request.user.profile.asociacion
            folder_id = request.POST.get('folder_id')

            if folder_id:
                client.post('drive/config', data={
                    'asociacion_id': asociacion.id,
                    'folder_id': folder_id
                })
                asociacion.drive_folder_id = folder_id
                asociacion.save()
                messages.success(request, 'Carpeta configurada correctamente')
        except Exception as e:
            messages.error(request, f'Error al configurar carpeta: {str(e)}')

    return redirect('/dashboard/backend/?tab=drive')



@association_required
def export_socias(request):
    """Exportar socias a CSV o Excel"""
    asociacion = request.user.profile.asociacion
    fmt = request.GET.get('format', 'csv')
    is_template = request.GET.get('template') == 'true'

    socias = Socia.objects.filter(asociacion=asociacion)

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        filename = 'plantilla_socias.csv' if is_template else 'socias.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        # Headers matching model fields
        headers = ['numero_socia', 'nombre', 'apellidos', 'telefono', 'email', 'direccion',
                   'numero', 'piso', 'escalera', 'codigo_postal', 'provincia', 'pais',
                   'nacimiento', 'pagado', 'descripcion']
        writer.writerow(headers)

        if not is_template:
            for socia in socias:
                writer.writerow([
                    socia.numero_socia, socia.nombre, socia.apellidos, socia.telefono, socia.email, socia.direccion,
                    socia.numero, socia.piso, socia.escalera, socia.codigo_postal, socia.provincia, socia.pais,
                    socia.nacimiento, socia.pagado, socia.descripcion
                ])

        return response

    elif fmt == 'excel':
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Socias"

            headers = ['numero_socia', 'nombre', 'apellidos', 'telefono', 'email', 'direccion',
                   'numero', 'piso', 'escalera', 'codigo_postal', 'provincia', 'pais',
                   'nacimiento', 'pagado', 'descripcion']
            ws.append(headers)

            if not is_template:
                for socia in socias:
                    ws.append([
                        socia.numero_socia, socia.nombre, socia.apellidos, socia.telefono, socia.email, socia.direccion,
                        socia.numero, socia.piso, socia.escalera, socia.codigo_postal, socia.provincia, socia.pais,
                        socia.nacimiento, socia.pagado, socia.descripcion
                    ])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'plantilla_socias.xlsx' if is_template else 'socias.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response

        except ImportError:
            messages.error(request, "Soporte para Excel no instalado. Descargando CSV.")
            return redirect(request.path + '?format=csv')

    return redirect('users:backend_management')


@association_required
def import_socias(request):
    """Importar socias desde CSV o Excel"""
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        asociacion = request.user.profile.asociacion

        try:
            import pandas as pd

            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            elif uploaded_file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(uploaded_file)
            else:
                messages.error(request, 'Formato no soportado. Usa CSV o Excel.')
                return redirect('users:backend_management')

            count, errors = _process_socias_import(df, asociacion)

            messages.success(request, f'Importación completada: {count} socias procesadas. {errors} errores.')

        except ImportError:
            messages.error(request, "Librería pandas no instalada.")
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')

    return redirect('users:backend_management')


@association_required
def delete_all_socias(request):
    """Eliminar todas las socias"""
    if request.method == 'POST':
        asociacion = request.user.profile.asociacion
        count = Socia.objects.filter(asociacion=asociacion).delete()[0]
        messages.success(request, f'Se han eliminado {count} socias correctamente.')

    return redirect('users:backend_management')

@association_required
def drive_callback(request):
    """Maneja el retorno de Google tras la autorización"""
    code = request.GET.get('code')
    error = request.GET.get('error')

    if error:
        messages.error(request, f'Error en autorización de Google: {error}')
        return redirect('users:backend_management')

    if code:
        client = get_client(request)
        asociacion = request.user.profile.asociacion
        try:
            client.post('drive/auth/callback', data={
                'code': code,
                'asociacion_id': asociacion.id
            })
            messages.success(request, 'Google Drive conectado correctamente')
        except Exception as e:
            messages.error(request, f'Error al conectar Drive: {str(e)}')

    return redirect('users:backend_management')


@association_required
def export_global_excel(request):
    """Exportar TODOS los datos a un solo Excel con múltiples hojas"""
    asociacion = request.user.profile.asociacion

    try:
        import openpyxl
        wb = openpyxl.Workbook()

        # --- HOJA 1: SOCIAS ---
        ws_socias = wb.active
        ws_socias.title = "Socias"
        headers_socias = ['numero_socia', 'nombre', 'apellidos', 'telefono', 'email', 'direccion',
                   'numero', 'piso', 'escalera', 'codigo_postal', 'provincia', 'pais',
                   'nacimiento', 'pagado', 'descripcion']
        ws_socias.append(headers_socias)

        for socia in Socia.objects.filter(asociacion=asociacion):
            ws_socias.append([
                socia.numero_socia, socia.nombre, socia.apellidos, socia.telefono, socia.email, socia.direccion,
                socia.numero, socia.piso, socia.escalera, socia.codigo_postal, socia.provincia, socia.pais,
                socia.nacimiento, socia.pagado, socia.descripcion
            ])

        # --- HOJA 2: LUGARES ---
        ws_lugares = wb.create_sheet("Lugares")
        headers_lugares = ['nombre', 'direccion', 'descripcion', 'numero', 'cp', 'ciudad', 'pais']
        ws_lugares.append(headers_lugares)

        for lugar in Lugar.objects.filter(asociacion=asociacion):
            ws_lugares.append([
                lugar.nombre, lugar.direccion, lugar.descripcion,
                lugar.numero, lugar.cp, lugar.ciudad, lugar.pais
            ])

        # --- HOJA 3: PERSONAS ---
        ws_personas = wb.create_sheet("Personas")
        headers_personas = ['nombre', 'apellidos', 'contacto', 'cargo', 'telefono', 'email', 'observaciones', 'proyecto_nombre']
        ws_personas.append(headers_personas)

        for persona in Persona.objects.filter(asociacion=asociacion):
            ws_personas.append([
                persona.nombre, persona.apellidos, persona.contacto, persona.cargo,
                persona.telefono, persona.email, persona.observaciones,
                persona.proyecto.nombre if persona.proyecto else ''
            ])

        # --- HOJA 4: MATERIALES ---
        ws_materiales = wb.create_sheet("Materiales")
        headers_materiales = ['nombre', 'uso', 'precio', 'lugar_nombre', 'encargado_persona_nombre', 'encargado_socia_numero']
        ws_materiales.append(headers_materiales)

        for material in Material.objects.filter(asociacion=asociacion):
            ws_materiales.append([
                material.nombre, material.uso, material.precio,
                material.lugar.nombre if material.lugar else '',
                f"{material.encargado_persona.nombre} {material.encargado_persona.apellidos}" if material.encargado_persona else '',
                material.encargado_socia.numero_socia if material.encargado_socia else ''
            ])

        # --- HOJA 5: CONTABILIDAD ---
        ws_finanzas = wb.create_sheet("Contabilidad")
        headers_finanzas = ['fecha_transaccion', 'cantidad', 'concepto', 'descripcion', 'entidad', 'fecha_vencimiento']
        ws_finanzas.append(headers_finanzas)

        for trans in Transaccion.objects.filter(asociacion=asociacion):
            ws_finanzas.append([
                trans.fecha_transaccion, trans.cantidad, trans.concepto, trans.descripcion,
                trans.entidad, trans.fecha_vencimiento
            ])

        # --- HOJA 6: ACTIVIDADES ---
        ws_eventos = wb.create_sheet("Actividades")
        headers_eventos = ['nombre', 'fecha', 'lugar', 'responsable_numero_socia', 'responsable_nombre',
                           'descripcion', 'duracion', 'colaboradores', 'observaciones']
        ws_eventos.append(headers_eventos)

        for evento in Evento.objects.filter(asociacion=asociacion):
            # Excel no soporta timezones
            fecha_naive = evento.fecha.replace(tzinfo=None) if evento.fecha else None

            # Obtener nombre del lugar (FK o texto)
            lugar_str = str(evento.lugar) if evento.lugar else evento.lugar_nombre

            ws_eventos.append([
                evento.nombre, fecha_naive, lugar_str,
                evento.responsable.numero_socia if evento.responsable else '',
                str(evento.responsable) if evento.responsable else '',
                evento.descripcion, evento.duracion, evento.colaboradores,
                evento.observaciones
            ])

        # --- HOJA 7: PROYECTOS ---
        ws_proyectos = wb.create_sheet("Proyectos")
        headers_proyectos = ['nombre', 'responsable', 'fecha_inicio', 'fecha_final', 'lugar',
                             'descripcion', 'materiales', 'involucrados', 'recursivo']
        ws_proyectos.append(headers_proyectos)

        for proy in Proyecto.objects.filter(asociacion=asociacion):
            # Obtener lugar (FK o texto)
            lugar_str = str(proy.lugar_fk) if proy.lugar_fk else proy.lugar

            ws_proyectos.append([
                proy.nombre,
                str(proy.responsable) if proy.responsable else '',
                proy.fecha_inicio, proy.fecha_final, lugar_str,
                proy.descripcion, proy.materiales, proy.involucrados, proy.recursivo
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="backup_completo_{asociacion.nombre}.xlsx"'
        wb.save(response)
        return response

    except ImportError:
        messages.error(request, "Soporte para Excel no instalado.")
        return redirect('users:backend_management')
    except Exception as e:
        messages.error(request, f"Error al generar Excel global: {str(e)}")
        return redirect('users:backend_management')


@association_required
def import_global_excel(request):
    """Importar TODOS los datos desde un solo Excel"""
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        asociacion = request.user.profile.asociacion

        try:
            import pandas as pd

            if not uploaded_file.name.endswith(('.xls', '.xlsx')):
                messages.error(request, 'Formato no soportado. Usa Excel (.xlsx).')
                return redirect('users:backend_management')

            # Leer todas las hojas
            xls = pd.read_excel(uploaded_file, sheet_name=None)

            results = []

            # 1. Procesar Socias (Prioridad alta por dependencias)
            if 'Socias' in xls:
                count, errors = _process_socias_import(xls['Socias'], asociacion)
                results.append(f"Socias: {count} importadas, {errors} errores.")

            # 2. Procesar Lugares
            if 'Lugares' in xls:
                count = _process_lugares_import(xls['Lugares'], asociacion)
                results.append(f"Lugares: {count} importados.")

            # 3. Procesar Personas
            if 'Personas' in xls:
                count = _process_personas_import(xls['Personas'], asociacion)
                results.append(f"Personas: {count} importadas.")

            # 4. Procesar Materiales
            if 'Materiales' in xls:
                count = _process_materiales_import(xls['Materiales'], asociacion)
                results.append(f"Materiales: {count} importados.")

            # 5. Procesar Contabilidad
            if 'Contabilidad' in xls:
                count = _process_finanzas_import(xls['Contabilidad'], asociacion)
                results.append(f"Contabilidad: {count} importadas.")

            # 6. Procesar Proyectos
            if 'Proyectos' in xls:
                count = _process_proyectos_import(xls['Proyectos'], asociacion)
                results.append(f"Proyectos: {count} importados.")

            # 7. Procesar Actividades (Depende de Socias)
            if 'Actividades' in xls:
                count = _process_eventos_import(xls['Actividades'], asociacion)
                results.append(f"Actividades: {count} importadas.")

            if not results:
                messages.warning(request, "El archivo no contiene ninguna hoja válida.")
            else:
                messages.success(request, "Importación Global Completa: " + " | ".join(results))

        except ImportError:
            messages.error(request, "Librería pandas no instalada.")
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo global: {str(e)}')

    return redirect('users:backend_management')


# --- HELPERS DE IMPORTACIÓN ---

def _process_socias_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    errors = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('numero_socia')) or pd.isna(row.get('nombre')): continue

            def get_val(col):
                val = row.get(col)
                return str(val) if not pd.isna(val) else ''

            Socia.objects.update_or_create(
                asociacion=asociacion,
                numero_socia=str(row['numero_socia']),
                defaults={
                    'nombre': get_val('nombre'),
                    'apellidos': get_val('apellidos'),
                    'telefono': get_val('telefono'),
                    'email': get_val('email'),
                    'direccion': get_val('direccion'),
                    'numero': get_val('numero'),
                    'piso': get_val('piso'),
                    'escalera': get_val('escalera'),
                    'codigo_postal': get_val('codigo_postal'),
                    'provincia': get_val('provincia'),
                    'pais': get_val('pais') or 'España',
                    'pagado': str(row.get('pagado', '')).lower() in ['true', '1', 'si', 'yes'],
                    'descripcion': get_val('descripcion')
                }
            )
            count += 1
        except Exception:
            errors += 1
    return count, errors

def _process_lugares_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('nombre')): continue

            Lugar.objects.update_or_create(
                asociacion=asociacion,
                nombre=str(row['nombre']),
                defaults={
                    'direccion': str(row.get('direccion', '')),
                    'descripcion': str(row.get('descripcion', '')),
                    'numero': str(row.get('numero', '')),
                    'cp': str(row.get('cp', '')),
                    'ciudad': str(row.get('ciudad', '')),
                    'pais': str(row.get('pais', 'España'))
                }
            )
            count += 1
        except Exception: pass
    return count

def _process_personas_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('nombre')): continue

            # Buscar proyecto si existe
            proyecto = None
            if not pd.isna(row.get('proyecto_nombre')):
                proyecto = Proyecto.objects.filter(asociacion=asociacion, nombre=str(row['proyecto_nombre'])).first()

            Persona.objects.update_or_create(
                asociacion=asociacion,
                nombre=str(row['nombre']),
                apellidos=str(row.get('apellidos', '')),
                defaults={
                    'contacto': str(row.get('contacto', '')),
                    'cargo': str(row.get('cargo', '')),
                    'telefono': str(row.get('telefono', '')),
                    'email': str(row.get('email', '')),
                    'observaciones': str(row.get('observaciones', '')),
                    'proyecto': proyecto
                }
            )
            count += 1
        except Exception: pass
    return count

def _process_materiales_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('nombre')): continue

            # Buscar lugar
            lugar = None
            if not pd.isna(row.get('lugar_nombre')):
                lugar = Lugar.objects.filter(asociacion=asociacion, nombre=str(row['lugar_nombre'])).first()

            # Buscar encargado socia
            encargado_socia = None
            if not pd.isna(row.get('encargado_socia_numero')):
                encargado_socia = Socia.objects.filter(asociacion=asociacion, numero_socia=str(row['encargado_socia_numero'])).first()

            # Buscar encargado persona (nombre completo aproximado)
            encargado_persona = None
            if not pd.isna(row.get('encargado_persona_nombre')):
                # Intento simple de búsqueda por nombre
                nombre_completo = str(row['encargado_persona_nombre'])
                encargado_persona = Persona.objects.filter(asociacion=asociacion, nombre__icontains=nombre_completo.split()[0]).first()

            Material.objects.update_or_create(
                asociacion=asociacion,
                nombre=str(row['nombre']),
                defaults={
                    'uso': str(row.get('uso', '')),
                    'precio': row.get('precio', 0.0),
                    'lugar': lugar,
                    'encargado_socia': encargado_socia,
                    'encargado_persona': encargado_persona
                }
            )
            count += 1
        except Exception: pass
    return count

def _process_finanzas_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('cantidad')) or pd.isna(row.get('fecha_transaccion')): continue

            Transaccion.objects.create(
                asociacion=asociacion,
                fecha_transaccion=pd.to_datetime(row['fecha_transaccion']).date(),
                cantidad=row['cantidad'],
                concepto=str(row.get('concepto', 'Importado')),
                descripcion=str(row.get('descripcion', '')),
                entidad=str(row.get('entidad', '')),
                fecha_vencimiento=pd.to_datetime(row['fecha_vencimiento']).date() if not pd.isna(row.get('fecha_vencimiento')) else None
            )
            count += 1
        except Exception: pass
    return count

def _process_eventos_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('nombre')) or pd.isna(row.get('fecha')): continue

            responsable = None
            if not pd.isna(row.get('responsable_numero_socia')):
                responsable = Socia.objects.filter(asociacion=asociacion, numero_socia=str(row['responsable_numero_socia'])).first()

            if not responsable:
                responsable = Socia.objects.filter(asociacion=asociacion).first()
                if not responsable: continue

            Evento.objects.create(
                asociacion=asociacion,
                nombre=str(row['nombre']),
                fecha=pd.to_datetime(row['fecha']),
                lugar=None, # Se asigna abajo si existe
                lugar_nombre=str(row.get('lugar', '')), # Guardar como texto por defecto
                responsable=responsable,
                descripcion=str(row.get('descripcion', '')),
                colaboradores=str(row.get('colaboradores', '')),
                observaciones=str(row.get('observaciones', ''))
            )

            # Intentar vincular lugar si existe
            if not pd.isna(row.get('lugar')):
                lugar_obj = Lugar.objects.filter(asociacion=asociacion, nombre=str(row['lugar'])).first()
                if lugar_obj:
                    evt = Evento.objects.filter(asociacion=asociacion, nombre=str(row['nombre'])).last()
                    evt.lugar = lugar_obj
                    evt.save()

            count += 1
        except Exception: pass
    return count

def _process_proyectos_import(df, asociacion):
    import pandas as pd
    df.columns = df.columns.str.lower()
    count = 0
    for _, row in df.iterrows():
        try:
            if pd.isna(row.get('nombre')) or pd.isna(row.get('fecha_inicio')): continue

            Proyecto.objects.create(
                asociacion=asociacion,
                nombre=str(row['nombre']),
                responsable=None, # Se asigna si es socia, pero aquí es texto en el modelo? No, es FK.
                # El modelo Proyecto tiene responsable como FK a Socia.
                # En el excel exportamos str(responsable).
                # Deberíamos intentar buscar la socia por nombre o algo, pero es difícil.
                # Por ahora lo dejamos null si no coincide exacto, o intentamos buscar.
                fecha_inicio=pd.to_datetime(row['fecha_inicio']).date(),
                fecha_final=pd.to_datetime(row['fecha_final']).date() if not pd.isna(row.get('fecha_final')) else None,
                lugar=str(row.get('lugar', '')),
                descripcion=str(row.get('descripcion', '')),
                materiales=str(row.get('materiales', '')),
                involucrados=str(row.get('involucrados', '')),
                recursivo=str(row.get('recursivo', '')).lower() in ['true', '1', 'si', 'yes']
            )

            # Intentar vincular lugar FK
            if not pd.isna(row.get('lugar')):
                lugar_obj = Lugar.objects.filter(asociacion=asociacion, nombre=str(row['lugar'])).first()
                if lugar_obj:
                    proy = Proyecto.objects.filter(asociacion=asociacion, nombre=str(row['nombre'])).last()
                    proy.lugar_fk = lugar_obj
                    proy.save()

            count += 1
        except Exception: pass
    return count


# --- GESTIÓN FINANZAS ---

@association_required
def export_finanzas(request):
    asociacion = request.user.profile.asociacion
    fmt = request.GET.get('format', 'csv')
    is_template = request.GET.get('template') == 'true'

    queryset = Transaccion.objects.filter(asociacion=asociacion)
    headers = ['fecha_transaccion', 'cantidad', 'concepto', 'descripcion', 'entidad', 'fecha_vencimiento']

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        filename = 'plantilla_contabilidad.csv' if is_template else 'contabilidad.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(headers)
        if not is_template:
            for obj in queryset:
                writer.writerow([obj.fecha_transaccion, obj.cantidad, obj.concepto, obj.descripcion, obj.entidad, obj.fecha_vencimiento])
        return response

    elif fmt == 'excel':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contabilidad"
            ws.append(headers)
            if not is_template:
                for obj in queryset:
                    ws.append([obj.fecha_transaccion, obj.cantidad, obj.concepto, obj.descripcion, obj.entidad, obj.fecha_vencimiento])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'plantilla_contabilidad.xlsx' if is_template else 'contabilidad.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        except ImportError:
            messages.error(request, "Soporte para Excel no instalado.")
            return redirect(request.path + '?format=csv')

    return redirect('users:backend_management')

@association_required
def import_finanzas(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        asociacion = request.user.profile.asociacion
        try:
            import pandas as pd
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            elif uploaded_file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(uploaded_file)
            else:
                messages.error(request, 'Formato no soportado.')
                return redirect('users:backend_management')

            count = _process_finanzas_import(df, asociacion)
            messages.success(request, f'Importadas {count} transacciones.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('users:backend_management')

@association_required
def delete_all_finanzas(request):
    if request.method == 'POST':
        count = Transaccion.objects.filter(asociacion=request.user.profile.asociacion).delete()[0]
        messages.success(request, f'Eliminadas {count} transacciones.')
    return redirect('users:backend_management')


# --- GESTIÓN ACTIVIDADES ---

@association_required
def export_eventos(request):
    asociacion = request.user.profile.asociacion
    fmt = request.GET.get('format', 'csv')
    is_template = request.GET.get('template') == 'true'

    queryset = Evento.objects.filter(asociacion=asociacion)
    headers = ['nombre', 'fecha', 'lugar', 'responsable_numero_socia', 'descripcion', 'duracion', 'colaboradores', 'observaciones']

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        filename = 'plantilla_actividades.csv' if is_template else 'actividades.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(headers)
        if not is_template:
            for obj in queryset:
                writer.writerow([
                    obj.nombre, obj.fecha, obj.lugar,
                    obj.responsable.numero_socia if obj.responsable else '',
                    obj.descripcion, obj.duracion, obj.colaboradores, obj.observaciones
                ])
        return response

    elif fmt == 'excel':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Actividades"
            ws.append(headers)
            if not is_template:
                for obj in queryset:
                    # Excel no soporta timezones
                    fecha_naive = obj.fecha.replace(tzinfo=None) if obj.fecha else None
                    # Obtener nombre del lugar (FK o texto)
                    lugar_str = str(obj.lugar) if obj.lugar else obj.lugar_nombre

                    ws.append([
                        obj.nombre, fecha_naive, lugar_str,
                        obj.responsable.numero_socia if obj.responsable else '',
                        obj.descripcion, obj.duracion, obj.colaboradores, obj.observaciones
                    ])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'plantilla_actividades.xlsx' if is_template else 'actividades.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        except ImportError:
            messages.error(request, "Soporte para Excel no instalado.")
            return redirect(request.path + '?format=csv')

    return redirect('users:backend_management')

@association_required
def import_eventos(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        asociacion = request.user.profile.asociacion
        try:
            import pandas as pd
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            elif uploaded_file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(uploaded_file)
            else:
                messages.error(request, 'Formato no soportado.')
                return redirect('users:backend_management')

            count = _process_eventos_import(df, asociacion)
            messages.success(request, f'Importadas {count} actividades.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('users:backend_management')

@association_required
def delete_all_eventos(request):
    if request.method == 'POST':
        count = Evento.objects.filter(asociacion=request.user.profile.asociacion).delete()[0]
        messages.success(request, f'Eliminadas {count} actividades.')
    return redirect('users:backend_management')


# --- GESTIÓN PROYECTOS ---

@association_required
def export_proyectos(request):
    asociacion = request.user.profile.asociacion
    fmt = request.GET.get('format', 'csv')
    is_template = request.GET.get('template') == 'true'

    queryset = Proyecto.objects.filter(asociacion=asociacion)
    headers = ['nombre', 'responsable', 'fecha_inicio', 'fecha_final', 'lugar', 'descripcion', 'materiales', 'involucrados', 'recursivo']

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        filename = 'plantilla_proyectos.csv' if is_template else 'proyectos.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(headers)
        if not is_template:
            for obj in queryset:
                writer.writerow([
                    obj.nombre, obj.responsable, obj.fecha_inicio, obj.fecha_final, obj.lugar,
                    obj.descripcion, obj.materiales, obj.involucrados, obj.recursivo
                ])
        return response

    elif fmt == 'excel':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Proyectos"
            ws.append(headers)
            if not is_template:
                for obj in queryset:
                    # Obtener lugar (FK o texto)
                    lugar_str = str(obj.lugar_fk) if obj.lugar_fk else obj.lugar

                    ws.append([
                        obj.nombre,
                        str(obj.responsable) if obj.responsable else '',
                        obj.fecha_inicio, obj.fecha_final, lugar_str,
                        obj.descripcion, obj.materiales, obj.involucrados, obj.recursivo
                    ])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'plantilla_proyectos.xlsx' if is_template else 'proyectos.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        except ImportError:
            messages.error(request, "Soporte para Excel no instalado.")
            return redirect(request.path + '?format=csv')

    return redirect('users:backend_management')

@association_required
def import_proyectos(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        asociacion = request.user.profile.asociacion
        try:
            import pandas as pd
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            elif uploaded_file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(uploaded_file)
            else:
                messages.error(request, 'Formato no soportado.')
                return redirect('users:backend_management')

            count = _process_proyectos_import(df, asociacion)
            messages.success(request, f'Importados {count} proyectos.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('users:backend_management')

@association_required
def delete_all_proyectos(request):
    if request.method == 'POST':
        count = Proyecto.objects.filter(asociacion=request.user.profile.asociacion).delete()[0]
        messages.success(request, f'Eliminados {count} proyectos.')
    return redirect('users:backend_management')